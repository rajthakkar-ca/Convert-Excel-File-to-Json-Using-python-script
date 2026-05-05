from __future__ import annotations

"""
Convert a clinic-scoreboard .xlsx into a queryable JSON document.

The scoreboard layout (one sheet, very wide, very few rows) does not fit into a
standard "first row is headers, the rest is data" model, so a naive
pandas/csv-style dump produces JSON that is technically lossless but painful to
work with.
"""

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_HEADER_ROWS = {
    "section": 1,
    "metric_name": 2,
    "focus": 3,
    "source": 4,
    "role": 5,
    "target_label": 6,
    "target": 7,
    "data_start": 8,
}

ROW_LABEL_COLUMN = 1

_slug_re = re.compile(r"[^a-z0-9]+")


def slugify(text: str) -> str:
    text = text.replace("\n", " ").replace("\r", " ").strip().lower()
    text = _slug_re.sub("_", text)
    return text.strip("_") or "unnamed"


def clean_header(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = " ".join(value.split())
        return cleaned or None
    return str(value)


def jsonable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        if isinstance(value, dt.datetime) and (
            value.hour or value.minute or value.second
        ):
            return value.isoformat()
        return (value.date() if isinstance(value, dt.datetime) else value).isoformat()
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def is_formula(cell_value: Any) -> bool:
    return isinstance(cell_value, str) and cell_value.startswith("=")


def cell_pair(formula_ws: Worksheet, value_ws: Worksheet, row: int, col: int) -> dict:
    raw = formula_ws.cell(row=row, column=col).value
    if raw is None:
        return {"value": None, "formula": None}
    if is_formula(raw):
        cached = value_ws.cell(row=row, column=col).value
        return {"value": jsonable(cached), "formula": raw}
    return {"value": jsonable(raw), "formula": None}


def section_headers(ws: Worksheet, section_row: int) -> dict[int, str]:
    out: dict[int, str] = {}
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= section_row <= merged.max_row:
            title = ws.cell(row=merged.min_row, column=merged.min_col).value
            title = clean_header(title)
            if title:
                for col in range(merged.min_col, merged.max_col + 1):
                    out[col] = title

    for col in range(1, ws.max_column + 1):
        if col in out:
            continue
        v = clean_header(ws.cell(row=section_row, column=col).value)
        if v:
            out[col] = v
    return out


def detect_groups(metric_columns: list[int]) -> list[list[int]]:
    groups: list[list[int]] = []
    current: list[int] = []
    for col in metric_columns:
        if not current or col == current[-1] + 1:
            current.append(col)
        else:
            groups.append(current)
            current = [col]
    if current:
        groups.append(current)
    return groups


def derive_group_title(
    metric_names: list[str], section_titles: list[str | None]
) -> str | None:
    explicit = {t for t in section_titles if t}
    if len(explicit) == 1:
        return next(iter(explicit))

    if not metric_names:
        return None
    parts = [n.split() for n in metric_names if n]
    if not parts:
        return None

    prefix: list[str] = []
    for tokens in zip(*parts):
        if all(t == tokens[0] for t in tokens):
            prefix.append(tokens[0])
        else:
            break

    if prefix and len(prefix) < len(parts[0]):
        return " ".join(prefix)
    return None


def convert(
    xlsx_path: Path,
    sheet_name: str | None = None,
    header_rows: dict[str, int] | None = None,
) -> dict[str, Any]:

    rows = {**DEFAULT_HEADER_ROWS, **(header_rows or {})}

    wb_f = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_v = openpyxl.load_workbook(xlsx_path, data_only=True)

    sheet_name = sheet_name or wb_f.sheetnames[0]
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    section_map = section_headers(ws_f, rows["section"])

    metric_columns: list[int] = []
    raw_records: list[dict[str, Any]] = []

    for col in range(1, ws_f.max_column + 1):
        if col == ROW_LABEL_COLUMN:
            continue

        raw_name = ws_f.cell(row=rows["metric_name"], column=col).value
        name = clean_header(raw_name)

        if not name or name in {"\\"}:
            continue

        metric_columns.append(col)

        raw_records.append(
            {
                "column": get_column_letter(col),
                "column_index": col,
                "name": name,
                "raw_name": raw_name if raw_name != name else None,
                "section_title": section_map.get(col),
                "focus": clean_header(ws_f.cell(row=rows["focus"], column=col).value),
                "source": clean_header(ws_f.cell(row=rows["source"], column=col).value),
                "role": clean_header(ws_f.cell(row=rows["role"], column=col).value),
                "target": cell_pair(ws_f, ws_v, rows["target"], col),
            }
        )

    groups_cols = detect_groups(metric_columns)
    col_to_group: dict[int, int] = {}
    groups_out: list[dict[str, Any]] = []

    for idx, cols in enumerate(groups_cols):
        records_in_group = [r for r in raw_records if r["column_index"] in cols]

        title = derive_group_title(
            [r["name"] for r in records_in_group],
            [r["section_title"] for r in records_in_group],
        )

        for c in cols:
            col_to_group[c] = idx

        groups_out.append(
            {
                "index": idx,
                "title": title,
                "columns": [get_column_letter(c) for c in cols],
            }
        )

    # Stable, unique metric ids — collisions get a group-prefix or column suffix.
    slug_to_records: dict[str, list[dict[str, Any]]] = {}
    for r in raw_records:
        slug_to_records.setdefault(slugify(r["name"]), []).append(r)

    metrics_out: list[dict[str, Any]] = []
    seen_ids: set[str] = set()

    for r in raw_records:
        base = slugify(r["name"])
        candidates = [base]

        group_title = groups_out[col_to_group[r["column_index"]]]["title"]

        if group_title:
            candidates.append(f"{slugify(group_title)}_{base}")

        candidates.append(f"{base}_{r['column'].lower()}")

        if len(slug_to_records[base]) == 1:
            metric_id = base
        else:
            metric_id = next(c for c in candidates[1:] if c not in seen_ids)

        i = 2
        while metric_id in seen_ids:
            metric_id = f"{base}_{i}"
            i += 1

        seen_ids.add(metric_id)

        metrics_out.append(
            {
                "id": metric_id,
                "name": r["name"],
                "raw_name": r["raw_name"],
                "column": r["column"],
                "column_index": r["column_index"],
                "group_index": col_to_group[r["column_index"]],
                "section_title": r["section_title"],
                "focus": r["focus"],
                "source": r["source"],
                "role": r["role"],
                "target": r["target"],
            }
        )

    # Backfill metric_ids onto each group for easy lookup.
    for g in groups_out:
        g["metric_ids"] = [
            m["id"] for m in metrics_out if m["group_index"] == g["index"]
        ]

    # Weekly snapshots.
    weeks_out: list[dict[str, Any]] = []

    for row in range(rows["data_start"], ws_f.max_row + 1):
        label = ws_f.cell(row=row, column=ROW_LABEL_COLUMN).value

        if label is None:
            continue

        is_date = isinstance(label, (dt.date, dt.datetime))

        week = {
            "row": row,
            "week_ending": jsonable(label) if is_date else None,
            "label": jsonable(label) if not is_date else None,
            "values": {},
        }

        for m in metrics_out:
            week["values"][m["id"]] = cell_pair(
                ws_f, ws_v, row, m["column_index"]
            )

        weeks_out.append(week)

    # Preserve everything in column A (Focus:/Source:/Role:/dates and oddities like A6="...+", A7=6).
    row_labels: dict[int, Any] = {}
    for row in range(1, ws_f.max_row + 1):
        v = ws_f.cell(row=row, column=ROW_LABEL_COLUMN).value
        if v not in (None, ""):
            row_labels[row] = jsonable(v)

    return {
        "source_file": xlsx_path.name,
        "sheet_name": sheet_name,
        "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "header_rows": rows,
        "row_labels_column_a": row_labels,
        "groups": groups_out,
        "metrics": metrics_out,
        "weeks": weeks_out,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("-o", "--output", type=Path, default=Path("output.json"))
    parser.add_argument("--sheet", default=None)
    args = parser.parse_args()

    if not args.input.exists():
        print(f"error: input file not found: {args.input}", file=sys.stderr)
        return 2

    data = convert(args.input, sheet_name=args.sheet)
    args.output.write_text(json.dumps(data, indent=2, ensure_ascii=False))
    print(
        f"Wrote {args.output} "
        f"({len(data['metrics'])} metrics, "
        f"{len(data['groups'])} groups, "
        f"{len(data['weeks'])} weekly snapshots)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())